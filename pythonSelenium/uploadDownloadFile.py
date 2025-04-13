import time

import openpyxl.workbook
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support import expected_conditions as ec
from selenium.webdriver.support.wait import WebDriverWait


driver = webdriver.Chrome()  # If you want to use Firefox (GeckoDriver setup needed)
driver.implicitly_wait(5)
driver.get("https://rahulshettyacademy.com/upload-download-test/index.html")

driver.maximize_window()

print(driver.title)
upload_path = r"D:\Data\download.xlsx"  # Use a raw string for the file path

#Method to edit excel dynamically after download
def update_excel_data(filePath, searchTerm, colName, new_value):
    book = openpyxl.load_workbook(filePath)
    sheet = book.active
    Dict = {}

    for i in range (1, sheet.max_column+1): #find price column
        if sheet.cell(row=1, column=i).value ==colName:
            Dict['col']=i

    for j in range (1, sheet.max_row+1): #find apple
        for k in range (1, sheet.max_column+1):
            if sheet.cell(row = j ,column=k ).value == searchTerm:
                Dict['row']=j


    #put these found rows and col in the Dict
    sheet.cell(row = Dict['row'], column = Dict['col']).value == new_value
    book.save(upload_path)

#call the method.
update_excel_data(upload_path, 'Apple', 'price', 350)

upload_button = driver.find_element(By.CSS_SELECTOR,"input[type='file']")
upload_button.send_keys(upload_path)

wait = WebDriverWait(driver , 6)
toast_locator = (By.CSS_SELECTOR,"div[role='alert'] div:nth-child(2)")
wait.until(ec.visibility_of_element_located(toast_locator))
print(driver.find_element(*toast_locator).text)


fruit_name = driver.find_element(By.CSS_SELECTOR,'div[id="row-1"]>div:nth-child(2)').text

actual_price = driver.find_element(By.XPATH,'//div[@id="row-1"]/div[@id="cell-4-undefined"]').text

print(actual_price) # CSS : div[id='row-1']>div:nth-child(4)
print(fruit_name)


