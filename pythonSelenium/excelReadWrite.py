import openpyxl
#open workbook
book = openpyxl.load_workbook("D:\\Data\\Study\\Learn Python\\pythonAutomation\\pythonSelenium\\Test.xlsx")

#take control to active sheet
sheet = book.active

#take control to cell that needs to be printed
print(sheet.cell(row=2, column=2).value)

#sheet.cell(['C3']).value = 'Agg'

#Write the value in sheet
sheet.cell(row=3,column=3).value = 'Aggarwal'
print(sheet.cell(row=3,column=3).value)

print("rows = " , sheet.max_row)
print("columns =" , sheet.max_column)

#pick directly the cell number from sheet and print
print(sheet['B4'].value)

#print cols and row
for i in range (1, sheet.max_row+1):
 for j in range (1,sheet.max_column+1):
     print(sheet.cell(row=i,column=j).value)

#print only specific row

for i in range (1,sheet.max_row+1):
    if sheet.cell(row=i,column=1).value=='T2':
        for j in range(1, sheet.max_column + 1):
            print(sheet.cell(row=i, column=j).value)